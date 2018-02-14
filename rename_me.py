import os.path
import argparse
import zipfile
import os
from openpyxl import load_workbook
import shutil

class RENamer():
    parsedargs = dict()
    def __init__(self):
        parser = argparse.ArgumentParser(description='renamer is a script to rename statements')
        parser.add_argument('--template', dest="template", required=True, help="Template file path to look up account names")
        parser.add_argument('--outputdir', dest="outputdir", required=True, help="Output directory is where renamed files will be stored")
        parser.add_argument('--zipfile', dest="zipfile", required=True, help="Zipfile path which contains the monthly statements")
        global parsedargs
        parsedargs = parser.parse_args()

    def validate(self):
        global parsedargs
        if not os.path.isfile(parsedargs.zipfile) or not parsedargs.zipfile.endswith(".zip"):
             print '{0} file not found or is not a .zip file'.format(parsedargs.zipfile)
             exit(-1)
        if not os.path.isdir(parsedargs.outputdir):
         print 'output directory not found at {0} '.format(parsedargs.outputdir)
         exit(-1)
        if not os.path.isfile(parsedargs.template) or not (parsedargs.template.endswith(".xlsx") or parsedargs.template.endswith(".xls")):
         print 'template file not found at {0} '.format(parsedargs.template)
         exit(-1)

    def unzip_billing_statements(self):
        global parsedargs
        zip_ref = zipfile.ZipFile(parsedargs.zipfile, 'r')
        zip_ref.extractall(parsedargs.zipfile[:-4])
        zip_ref.close()
        print 'Extracted billing statemets from {0} to {1}\n'.format(parsedargs.zipfile, parsedargs.zipfile[:-4])

    def rename_billing_statements(self):
        global parsedargs
        accmap = renamer.fetch_account_mapping()
        renamed_file_count = 0
        total_file_count = 0
        skipped_file_count = 0
        for root, dirs, files in os.walk(parsedargs.zipfile[:-4]):
            for billing_stmt in files:
                total_file_count += 1
                accno = str(billing_stmt).split("-")[3].lstrip('0')
                if accno not in accmap:
                    skipped_file_count += 1
                    print 'Account {0} not found in template file, skipping {1}\n'.format(accno, billing_stmt)
                    continue
                shutil.copy(os.path.join(parsedargs.zipfile[:-4], str(billing_stmt)), os.path.join(parsedargs.outputdir, accmap[accno]+".pdf"))
                renamed_file_count += 1

        print '**** Summary ****\n' \
              'Billing statements zip file {0}\n' \
              'Renaming Template {1}\n' \
              'Target folder {2}\n' \
              'Files to rename {3}\n' \
              'Successfully renamed {4}\n' \
              'Files for which account could not be found {5}\n' \
              '**** End of Summary ****' \
              ''.format(parsedargs.zipfile, parsedargs.template, parsedargs.outputdir, total_file_count, renamed_file_count, skipped_file_count)

    def fetch_account_mapping(self):
        global parsedargs
        accnos_names_map = dict()
        wb = load_workbook(parsedargs.template, data_only=True)
        ws = wb['Renaming Template']
        # validate ws here
        for row in range(2, ws.max_row):
            accno = str(ws["A"+str(row)].value)
            name = str(ws["D"+str(row)].value)
            accnos_names_map[accno] = name
        return accnos_names_map


renamer = RENamer()
renamer.validate()
renamer.unzip_billing_statements()
renamer.rename_billing_statements()

